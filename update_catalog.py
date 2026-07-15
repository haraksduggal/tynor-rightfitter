"""
update_catalog.py — Tynor Catalog Updater
==========================================
Run this script after editing catalog_master.xlsx to regenerate catalog.json.

Usage:
    python3 update_catalog.py

Place this file in the same folder as catalog_master.xlsx and catalog.json.
"""

import json, os, sys
from openpyxl import load_workbook

EXCEL_FILE = "catalog_master.xlsx"
OUTPUT_FILE = "catalog.json"

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: {EXCEL_FILE} not found in current folder.")
        print("Make sure you save the Excel file as 'catalog_master.xlsx' in this folder.")
        sys.exit(1)

    print(f"Reading {EXCEL_FILE}...")
    wb = load_workbook(EXCEL_FILE, data_only=True)

    if "Catalog" not in wb.sheetnames:
        print("ERROR: 'Catalog' sheet not found in the Excel file.")
        sys.exit(1)

    ws = wb["Catalog"]

    # Read headers from row 4
    headers = [ws.cell(row=4, column=ci).value for ci in range(1, ws.max_column + 1)]
    headers = [h for h in headers if h]

    products = []
    errors = []

    for ri in range(5, ws.max_row + 1):
        row = [ws.cell(row=ri, column=ci).value for ci in range(1, len(headers) + 1)]

        # Skip empty rows
        if not any(row):
            continue

        product_id = row[0]
        name       = row[1]
        brand      = row[2]
        body_part  = row[3]
        sub_cat    = row[4]
        short_desc = row[5] or ""
        rating     = row[6]
        image_url  = row[7] or ""

        # Validate required fields
        if not name:
            errors.append(f"Row {ri}: Missing product name")
            continue
        if brand not in ("Cure", "Sport", "Life"):
            errors.append(f"Row {ri} ({name}): Brand must be Cure, Sport, or Life — got '{brand}'")
            continue
        try:
            rating = float(rating) if rating else 4.0
        except (ValueError, TypeError):
            errors.append(f"Row {ri} ({name}): Invalid rating '{rating}' — using 4.0")
            rating = 4.0

        # Build prices dict
        prices = {}
        size_map = [
            (8,  "Small"),
            (9,  "Medium"),
            (10, "Large"),
            (11, "XL"),
            (12, "XXL"),
            (13, "One Size"),
        ]
        for col_idx, size_name in size_map:
            val = row[col_idx] if col_idx < len(row) else None
            if val is not None and str(val).strip():
                try:
                    prices[size_name] = int(float(str(val).replace(",", "").replace("₹", "").strip()))
                except ValueError:
                    errors.append(f"Row {ri} ({name}): Invalid price for {size_name}: '{val}'")

        # Build availability (in_stock for all available sizes)
        availability = {size: "in_stock" for size in prices}

        product = {
            "product_id":   str(product_id) if product_id else f"P{ri-4:03d}",
            "name":         str(name).strip(),
            "brand":        str(brand).strip(),
            "body_part":    str(body_part).strip() if body_part else "",
            "sub_cat":      str(sub_cat).strip() if sub_cat else "",
            "short_desc":   str(short_desc).strip(),
            "rating":       rating,
            "image_url":    str(image_url).strip(),
            "prices":       prices,
            "availability": availability,
            "activity":     [],
            "attributes":   [],
            "conditions":   [],
            "easy_names":   [],
            "is_synthetic": False,
        }
        products.append(product)

    if errors:
        print(f"\n⚠️  {len(errors)} warnings found:")
        for e in errors:
            print(f"   - {e}")

    if not products:
        print("ERROR: No products found. Check the Catalog sheet.")
        sys.exit(1)

    with open(OUTPUT_FILE, 'w') as f:
        json.dump(products, f, indent=2, ensure_ascii=False)

    print(f"\n✅ Done! {len(products)} products written to {OUTPUT_FILE}")
    print("Restart your Streamlit app to see the changes.")

if __name__ == "__main__":
    main()
